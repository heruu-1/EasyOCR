"""
EasyOCR Production Backend
OCR API dengan dukungan preview, save, history, dan export XLSX
"""

import os
import sys
import logging
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from io import BytesIO
import tempfile

# Import custom modules
from bukti_setor.processor import BuktiSetorProcessor
from utils.file_utils import allowed_file, cleanup_temp_files
from utils.helpers import validate_file_size, format_response

# Database setup
db = SQLAlchemy()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def create_app():
    """Application factory"""
    app = Flask(__name__)
    
    # Configuration
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
    app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
    
    # Database configuration
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        # Fix PostgreSQL URL for SQLAlchemy
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    else:
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bukti_setor.db'
    
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 300,
    }
    
    # Initialize extensions
    db.init_app(app)
    CORS(app, resources={
        r"/api/*": {"origins": "*", "methods": ["GET", "POST", "DELETE", "OPTIONS"]}
    })
    
    # Import models after db initialization
    from models import BuktiSetor
    
    # Initialize OCR processor
    processor = BuktiSetorProcessor()
    
    # Routes
    @app.route('/health', methods=['GET'])
    def health_check():
        """Health check endpoint"""
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'service': 'EasyOCR Backend'
        }), 200
    
    @app.route('/api/bukti_setor/process', methods=['POST'])
    def process_bukti_setor():
        """Process bukti setor with OCR"""
        try:
            if 'file' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            if not allowed_file(file.filename):
                return jsonify({'error': 'File type not allowed'}), 400
            
            # Validate file size
            if not validate_file_size(file):
                return jsonify({'error': 'File too large'}), 400
            
            # Process with temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
                file.save(temp_file.name)
                
                try:
                    # Process OCR
                    result = processor.process_file(temp_file.name)
                    
                    if result.get('success'):
                        return jsonify(format_response(result, 'OCR processing completed successfully'))
                    else:
                        return jsonify({'error': result.get('error', 'Processing failed')}), 500
                        
                except Exception as e:
                    logger.error(f"OCR processing error: {e}")
                    return jsonify({'error': f'Processing failed: {str(e)}'}), 500
                finally:
                    # Cleanup
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                        
        except Exception as e:
            logger.error(f"Request processing error: {e}")
            return jsonify({'error': 'Internal server error'}), 500
    
    @app.route('/api/bukti_setor/save', methods=['POST'])
    def save_bukti_setor():
        """Save OCR result to database"""
        try:
            data = request.get_json()
            
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            
            # Create new record
            bukti_setor = BuktiSetor(
                filename=data.get('filename', ''),
                ntpn=data.get('ntpn', ''),
                npwp=data.get('npwp', ''),
                nama_wajib_pajak=data.get('nama_wajib_pajak', ''),
                tanggal_setor=data.get('tanggal_setor', ''),
                jumlah_setor=data.get('jumlah_setor', 0),
                kode_akun_pajak=data.get('kode_akun_pajak', ''),
                kode_jenis_setoran=data.get('kode_jenis_setoran', ''),
                masa_pajak=data.get('masa_pajak', ''),
                tahun_pajak=data.get('tahun_pajak', ''),
                ocr_confidence=data.get('ocr_confidence', 0.0),
                raw_text=data.get('raw_text', ''),
                processing_time=data.get('processing_time', 0.0)
            )
            
            db.session.add(bukti_setor)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Data saved successfully',
                'id': bukti_setor.id,
                'data': bukti_setor.to_dict()
            }), 201
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Save error: {e}")
            return jsonify({'error': f'Failed to save data: {str(e)}'}), 500
    
    @app.route('/api/bukti_setor/history', methods=['GET'])
    def get_history():
        """Get processing history"""
        try:
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 10, type=int)
            
            records = BuktiSetor.query.order_by(BuktiSetor.created_at.desc()).paginate(
                page=page, per_page=per_page, error_out=False
            )
            
            return jsonify({
                'success': True,
                'data': [record.to_dict() for record in records.items],
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': records.total,
                    'pages': records.pages,
                    'has_next': records.has_next,
                    'has_prev': records.has_prev
                }
            }), 200
            
        except Exception as e:
            logger.error(f"History retrieval error: {e}")
            return jsonify({'error': f'Failed to retrieve history: {str(e)}'}), 500
    
    @app.route('/api/bukti_setor/delete/<int:record_id>', methods=['DELETE'])
    def delete_record(record_id):
        """Delete a record"""
        try:
            record = BuktiSetor.query.get_or_404(record_id)
            db.session.delete(record)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Record deleted successfully'
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Delete error: {e}")
            return jsonify({'error': f'Failed to delete record: {str(e)}'}), 500
    
    @app.route('/api/bukti_setor/export', methods=['GET'])
    def export_to_xlsx():
        """Export records to XLSX"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            # Get all records
            records = BuktiSetor.query.order_by(BuktiSetor.created_at.desc()).all()
            
            if not records:
                return jsonify({'error': 'No data to export'}), 404
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bukti Setor Data"
            
            # Headers
            headers = [
                'ID', 'Filename', 'NTPN', 'NPWP', 'Nama Wajib Pajak',
                'Tanggal Setor', 'Jumlah Setor', 'Kode Akun Pajak',
                'Kode Jenis Setoran', 'Masa Pajak', 'Tahun Pajak',
                'OCR Confidence', 'Processing Time', 'Created At'
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row, record in enumerate(records, 2):
                data = record.to_dict()
                ws.cell(row=row, column=1, value=data['id'])
                ws.cell(row=row, column=2, value=data['filename'])
                ws.cell(row=row, column=3, value=data['ntpn'])
                ws.cell(row=row, column=4, value=data['npwp'])
                ws.cell(row=row, column=5, value=data['nama_wajib_pajak'])
                ws.cell(row=row, column=6, value=data['tanggal_setor'])
                ws.cell(row=row, column=7, value=data['jumlah_setor'])
                ws.cell(row=row, column=8, value=data['kode_akun_pajak'])
                ws.cell(row=row, column=9, value=data['kode_jenis_setoran'])
                ws.cell(row=row, column=10, value=data['masa_pajak'])
                ws.cell(row=row, column=11, value=data['tahun_pajak'])
                ws.cell(row=row, column=12, value=data['ocr_confidence'])
                ws.cell(row=row, column=13, value=data['processing_time'])
                ws.cell(row=row, column=14, value=data['created_at'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to memory
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"bukti_setor_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            return jsonify({'error': f'Failed to export data: {str(e)}'}), 500
    
    @app.route('/uploads/<filename>')
    def uploaded_file(filename):
        """Serve uploaded files"""
        try:
            return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
        except Exception as e:
            logger.error(f"File serving error: {e}")
            return jsonify({'error': 'File not found'}), 404
    
    return app

# Create app instance
app = create_app()

if __name__ == "__main__":
    # Production configuration
    port = int(os.environ.get("PORT", 8000))
    
    logger.info("🚀 Starting EasyOCR Production Server")
    logger.info(f"   Port: {port}")
    logger.info(f"   Upload folder: {app.config.get('UPLOAD_FOLDER', 'uploads')}")
    logger.info(f"   Memory optimization: enabled")
    logger.info(f"   Database: {'enabled' if db else 'disabled'}")
    
    # Ensure upload folder exists
    upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)
    
    # Initialize database tables
    try:
        with app.app_context():
            db.create_all()
            logger.info("✅ Database tables created/verified")
    except Exception as e:
        logger.error(f"❌ Database initialization error: {e}")
    
    try:
        # Production mode - no debug
        app.run(host="0.0.0.0", port=port, debug=False)
    except Exception as e:
        logger.error(f"❌ Failed to start server: {e}")
        sys.exit(1)